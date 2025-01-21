import json
import random
import openpyxl
from openpyxl.reader.excel import load_workbook
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from position.constants import URL_TEST, USERNAME_LOCATER, USERNAME, PASSWORD_LOCATER, PASSWORD, LOGIN_BUTTON, BUTTON1, \
    BUTTON2
from sql.connect_sql import getpgsql


# 登录页面
def login(driver):
    driver.open(url=URL_TEST)
    try:
        # 输入账号和密码，点击登录
        driver.input(method=By.XPATH, locator=USERNAME_LOCATER, text=USERNAME)
        driver.input(method=By.XPATH, locator=PASSWORD_LOCATER, text=PASSWORD)
        driver.click(method=By.XPATH, locator=LOGIN_BUTTON)
        driver.sleep()
    except (NoSuchElementException, TimeoutError):
        print("Element not found")
    try:
        driver.click(method=By.XPATH, locator=BUTTON1)
        driver.sleep()
    except (NoSuchElementException, TimeoutError):
        driver.locate(method=By.XPATH, locator=BUTTON2)
        driver.refresh()


# 加载Excel文件
def load_test_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过标题行
        test_data.append(row)
    return test_data


# 判断方案创建成功
def judge_plan_create(plan_name):
    try:
        sql_template = f"select * from neptune_artifact_scan_plan where plan_name = %s and is_deleted = %s"
        results = getpgsql(sql_template, (plan_name, 'N'))
        if results:  # 检查是否有返回结果
            if results[0]['plan_name'] == plan_name:
                print(f"自动化创建扫描方案成功:{results[0][0]}")
            else:
                print("自动化创建扫描方案失败")
        else:
            print("自动化创建扫描方案失败：未找到该方案")
    except Exception as e:
        print(f"数据库连接失败：{e}")


# 判断方案删除成功
def judge_plan_delete(plan_name=None):
    sql_template = "select * from neptune_artifact_scan_plan where plan_name = %s AND is_deleted = %s"
    results = getpgsql(sql_template, (plan_name, 'Y'))
    try:
        if results:  # 检查是否有返回结果
            if results[0]['plan_name'] == plan_name:
                print(f"自动化删除扫描方案成功: {plan_name}")
            else:
                print("查询结果与预期不符")
            return results
        else:
            print("自动化删除扫描方案失败：未找到该方案")
    except Exception as e:
        print(f"数据库操作失败：{e}")


# 判断方案编辑成功
def judge_plan_edit(plan_name=None, params=None):
    sql_template = "select * from neptune_artifact_scan_plan where plan_name = %s AND is_deleted = %s"
    results = getpgsql(sql_template, (plan_name, 'N'))
    try:
        if results:  # 检查是否有返回结果
            if params in results[0]['white_cve']:
                print(f"自动化编辑扫描方案成功: {plan_name}")
            else:
                print("查询结果与预期不符")
            return results
        else:
            print("自动化编辑扫描方案失败：未找到该方案")
    except Exception as e:
        print(f"数据库操作失败：{e}")


# 判断任务
def judge_task(sql_template, task_name=None):
    results = getpgsql(sql_template,(task_name,))
    try:
        if results:
            if results[0]['task_name'] == task_name:
                print(f"自动化查询扫描任务成功:{task_name}")
            else:
                print("自动化查询扫描任务失败")
            return results
        else:
            print("自动化查询扫描任务失败：未找到该任务")
    except Exception as e:
        print(f"数据库连接失败：{e}")


# 获取固定长度随机数
def get_number(length):
    num = random.randint(10 ** (length - 1), 10 ** length - 1)
    return num


# 保存cookies
def save_cookies(driver):
    cookies = driver.get_cookies()
    with open('cookies.json', 'w') as f:
        f.write(json.dumps(cookies))


# 使用cookies
def load_cookies(driver):
    try:
        with open('cookies.json') as f:
            cookies = json.loads(f.read())
        # 遍历字典获取cookies信息进行添加
        for cookie in cookies:
            driver.add_cookie(cookie)
        else:
            # 刷新页面，清除缓存
            driver.reflesh()
    except:
        print("没有可用的cookies信息")


# 从 Excel 文件中读取指定工作表的数据
def read_excel_data(file_path, sheet_name):
    """
    :param file_path: Excel 文件的路径
    :param sheet_name: 工作表的名称
    :return: 包含所有行数据的列表，每行数据以元组形式存储
    """
    # 参数验证
    if not file_path:
        raise ValueError("文件路径不能为空")
    if not sheet_name:
        raise ValueError("工作表名称不能为空")
    # 使用openpyxl库的load_workbook函数加载Excel文件
    workbook = openpyxl.load_workbook(file_path,read_only=True)
    # 检查工作表是否存在
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在")
    # 从加载的工作簿中获取指定名称的工作表
    sheet = workbook[sheet_name]
    # 初始化一个空列表，用于存储读取的数据
    data = []
    # 使用sheet的iter_rows方法遍历工作表中的行
    # min_row=2表示从第二行开始遍历（假设第一行是标题行）
    # values_only=True表示只获取单元格的值，不获取单元格对象
    try:
        for row in sheet.iter_rows(min_row=2,values_only=True):
            data.append(row)
    except FileNotFoundError:
        raise
    except openpyxl.utils.exceptions.InvalidFileException:
        raise
    except Exception as e:
        raise
    finally:
        # 确保工作簿被正确关闭，检查变量 workbook 是否在当前作用域中存在，如果存在，关闭工作簿
        if 'workbook' in locals():  # 内置函数，返回当前作用域中的所有局部变量及其值的字典
            workbook.close()
    return data


def read_excel_data2(file_path, sheet_name):
    """
    读取 Excel 文件中的数据，并返回一个列表的列表（不包含列名）。
    对 'items' 列进行特殊处理，将其转换为列表。
    :param file_path: Excel 文件的路径
    :param sheet_name: 工作表的名称
    :return: 包含所有行数据的列表，每行数据以列表形式存储
    """
    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    # 读取第一行作为标题（用于识别 'items' 列）
    headers = [cell.value for cell in sheet[1]]
    # 初始化数据列表
    data = []
    # 从第二行开始读取数据
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = []
        for col_num, cell_value in enumerate(row):
            if headers[col_num] == 'items' and cell_value:  # 处理 'items' 列
                row_data.append([item.strip() for item in cell_value.split(',')])
            else:
                row_data.append(cell_value)  # 其他列直接存储值
        data.append(row_data)
    return data


def read_excel_data3(file_path, sheet_name):
    """
    读取 Excel 文件中的数据，并返回一个列表的列表（不包含列名）。
    对 'items' 列进行特殊处理，将其转换为列表。
    :param file_path: Excel 文件的路径
    :param sheet_name: 工作表的名称
    :return: 包含所有行数据的列表，每行数据以字典形式存储
    """
    # 读取Excel文件中的数据
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    # 读取第一行作为标题
    headers = []
    for cell in sheet[1]:           # 遍历第一行的每个单元格
        headers.append(cell.value)  # 将单元格的值添加到列表中
    data = []
    # 从第二行开始读取数据
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for col_num, cell_value in enumerate(row, start=1):  # 枚举列号和单元格值
            header = headers[col_num - 1]  # 获取对应的标题
            if header == 'items':
    #           # 处理 'items' 列，将逗号分隔的字符串转换为列表
                if cell_value:  # 检查 cell_value 是否为空
                    row_data[header] = []
                    for item in cell_value.split(','):  # 遍历按逗号分割后的每个元素
                        row_data[header].append(item.strip())  # 去除元素前后的空白字符并添加到列表中
            else:
                row_data[header] = cell_value  # 对于其他列，直接存储值
        data.append(row_data)
    return data