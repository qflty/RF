from openpyxl.reader.excel import load_workbook
from selenium.common import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from position.consants import URL_TEST, USERNAME_LOCATER, USERNAME, PASSWORD_LOCATER, PASSWORD, LOGIN_BUTTON, BUTTON1, \
    BUTTON2


# 登录页面
def login(driver):
    driver.open(url=URL_TEST)
    try:
        # 输入账号和密码，点击登录
        driver.input(method=By.ID, locator=USERNAME_LOCATER, text=USERNAME)
        driver.input(method=By.ID, locator=PASSWORD_LOCATER, text=PASSWORD)
        driver.click(method=By.ID, locator=LOGIN_BUTTON)
        driver.sleep()
    except (NoSuchElementException, TimeoutError):
        print("Element not found by ID")
    try:
        driver.click(method=By.XPATH, locator=BUTTON1)
        driver.sleep()
    except (NoSuchElementException, TimeoutError):
        driver.locate(method=By.XPATH, locator=BUTTON2)
        driver.refresh()


# 加载Excel文件
def load_test_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过标题行
        test_data.append(row)
    return test_data
